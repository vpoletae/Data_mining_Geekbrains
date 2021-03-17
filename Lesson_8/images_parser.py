# importing google_images_download module 
from google_images_download import google_images_download
import openpyxl as opxl
import os

file_name = 'to_import.xlsx'
path = os.path.dirname(os.path.abspath(__file__))

wb = opxl.load_workbook(os.path.join(path, file_name))
ws = wb.active
search_queries = []

for row in range(1, ws.max_row+1):
      cell_value = ws.cell(row=row, column=1).value
      search_queries.append(cell_value)
  
# creating object 
response = google_images_download.googleimagesdownload()  
  
def downloadimages(query): 
      # keywords is the search query 
      # format is the image file format 
      # limit is the number of images to be downloaded 
      # print urs is to print the image file url 
      # size is the image size which can 
      # be specified manually ("large, medium, icon") 
      # aspect ratio denotes the height width ratio 
      # of images to download. ("tall, square, wide, panoramic") 
      arguments = {"keywords": query, 
                 "format": "png",
                 "limit":9, 
                 "print_urls":True, 
                 "size": "medium", 
                 "aspect_ratio": "panoramic"} 
      try: 
            response.download(arguments) 

      # Handling File NotFound Error     
      except FileNotFoundError:  
            arguments = {"keywords": query, 
                     "format": "png", 
                     "limit":4, 
                     "print_urls":True,  
                     "size": "medium"} 
                       
        # Providing arguments for the searched query 
      try: 
            # Downloading the photos based 
            # on the given arguments 
            response.download(arguments)  
      except: 
            pass
  
# Driver Code 
for query in search_queries: 
      downloadimages(query)  
      print()  
